"""Fail-closed parsers for the approved W3 pseudonymous workbook profiles.

The parser returns immutable raw-row lineage and normalized preview rows.  It
does not open a database session and cannot apply business facts.
"""

from __future__ import annotations

import json
import math
import posixpath
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time
from enum import StrEnum
from hashlib import sha256
from io import BytesIO
from typing import Generic, TypeAlias, TypeVar
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.domains.staff.policies import normalize_phone_number
from app.domains.w1c.policies import normalize_certification_number

MAX_COMPRESSED_BYTES = 10 * 1024 * 1024
MAX_PACKAGE_ENTRIES = 512
MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_ENTRY_BYTES = 32 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200
MAX_WORKSHEET_ROWS = 20_000
MAX_WORKSHEET_COLUMNS = 64
KST = ZoneInfo("Asia/Seoul")
VISIT_SERVICE_TYPES = frozenset({"방문요양", "방문목욕"})
_HH_MM = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
_TIMESTAMP = re.compile(
    r"^\d{4}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01]) "
    r"(?:[01]\d|2[0-3]):[0-5]\d:[0-5]\d$"
)
_DRIVE_PATH = re.compile(r"^[A-Za-z]:")
_ACTIVE_DECLARATION_MARKERS = (
    "activex",
    "connections",
    "customui",
    "dialogsheet",
    "externallink",
    "macroenabled",
    "macrosheet",
    "oleobject",
    "querytable",
    "vbaproject",
)


@dataclass(frozen=True, slots=True)
class WorkbookProfile:
    profile_version: str
    source_type: str
    sheet_name: str
    headers: tuple[str, ...]


NHIS_SCHEDULE_PROFILE_V1 = WorkbookProfile(
    profile_version="nhis-schedule-xlsx-v1",
    source_type="NHIS_SCHEDULE",
    sheet_name="일정계획",
    headers=(
        "일자",
        "시작시간",
        "종료시간",
        "수급자명",
        "수급자\n인정번호",
        "요양보호사명",
        "생년월일",
        "요양보호사번호",
        "종사자구분",
        "가족여부",
        "가족관계",
        "서비스구분",
        "수가코드",
        "수가명",
        "수가",
    ),
)

RFID_PROFILE_V1 = WorkbookProfile(
    profile_version="rfid-xlsx-v1",
    source_type="RFID",
    sheet_name="실시간전송내용",
    headers=(
        "구분",
        "수급자성명",
        "인정번호",
        "요양요원",
        "핸드폰번호",
        "급여종류",
        "총시간",
        "시작시간",
        "종료시간",
        "사용여부",
    ),
)


class RfidEventState(StrEnum):
    COMPLETE = "COMPLETE"
    START_ONLY = "START_ONLY"


class WorkbookParseBlocked(ValueError):
    """A profile or input violation with a guaranteed zero business-write count."""

    business_write_count = 0

    def __init__(
        self,
        code: str,
        message: str,
        *,
        source_row_number: int | None = None,
        column_name: str | None = None,
    ) -> None:
        self.code = code
        self.source_row_number = source_row_number
        self.column_name = column_name
        location = ""
        if source_row_number is not None:
            location += f" row={source_row_number}"
        if column_name is not None:
            location += f" column={column_name}"
        super().__init__(f"{code}:{location} {message}".rstrip())


RawCellValue: TypeAlias = str | int | float | bool | date | datetime | time | None


@dataclass(frozen=True, slots=True)
class RawSourceRow:
    sheet_ref: str
    source_row_number: int
    raw_values: tuple[RawCellValue, ...]

    @property
    def physical_address(self) -> tuple[str, int]:
        return (self.sheet_ref, self.source_row_number)


@dataclass(frozen=True, slots=True)
class NhisScheduleRow:
    sheet_ref: str
    source_row_number: int
    service_date: date
    planned_start: time
    planned_end: time
    declared_minutes: int
    recipient_name: str
    recipient_certification_number: str
    staff_name: str
    staff_birth_date: date
    staff_external_number: str
    worker_category: str
    family_flag: str
    family_relationship: str | None
    service_category: str
    fee_code: str
    fee_name: str
    fee_amount: int
    occurrence_signature: str
    occurrence_ordinal: int

    @property
    def occurrence_identity(self) -> str:
        return f"{self.occurrence_signature}:{self.occurrence_ordinal}"


@dataclass(frozen=True, slots=True)
class RfidRow:
    sheet_ref: str
    source_row_number: int
    transmission_kind: str
    recipient_name: str
    recipient_certification_number: str
    staff_name: str
    staff_phone: str
    staff_phone_normalized: str
    service_category: str
    reference_minutes: int
    actual_start: datetime
    actual_end: datetime | None
    use_state: str
    event_state: RfidEventState
    occurrence_signature: str
    occurrence_ordinal: int

    @property
    def occurrence_identity(self) -> str:
        return f"{self.occurrence_signature}:{self.occurrence_ordinal}"

    @property
    def actual_seconds(self) -> int | None:
        if self.actual_end is None:
            return None
        return int((self.actual_end - self.actual_start).total_seconds())

    @property
    def end_display(self) -> str:
        if self.actual_end is None:
            return f"종료X · {self.actual_start:%H:%M}"
        return self.actual_end.strftime("%H:%M")


ParsedRow = TypeVar("ParsedRow", NhisScheduleRow, RfidRow)


@dataclass(frozen=True, slots=True)
class ParsedWorkbook(Generic[ParsedRow]):
    profile_version: str
    source_type: str
    content_digest: str
    sheet_ref: str
    raw_rows: tuple[RawSourceRow, ...]
    parsed_rows: tuple[ParsedRow, ...]
    target_rows: tuple[ParsedRow, ...]
    warning_codes: tuple[str, ...] = ()
    business_write_count: int = 0


def _blocked(
    code: str,
    message: str,
    *,
    row: int | None = None,
    column: str | None = None,
) -> WorkbookParseBlocked:
    return WorkbookParseBlocked(
        code,
        message,
        source_row_number=row,
        column_name=column,
    )


def _unsafe_package_path(name: str) -> bool:
    if not name or "\x00" in name or "\\" in name or name.startswith("/"):
        return True
    if _DRIVE_PATH.match(name):
        return True
    normalized = posixpath.normpath(name)
    return normalized == ".." or normalized.startswith("../")


def _is_active_content(name: str) -> bool:
    lowered = name.casefold()
    blocked_segments = (
        "/activex/",
        "/dialogsheets/",
        "/embeddings/",
        "/externallinks/",
        "/macrosheets/",
        "/oleobjects/",
        "/querytables/",
        "/customui/",
    )
    if lowered.endswith("/vbaproject.bin") or lowered.endswith("/connections.xml"):
        return True
    if any(segment in f"/{lowered}" for segment in blocked_segments):
        return True
    return lowered.endswith(".bin") and not lowered.startswith("xl/printersettings/")


def _declares_active_content(value: str) -> bool:
    lowered = value.casefold()
    return any(marker in lowered for marker in _ACTIVE_DECLARATION_MARKERS)


def _preflight_xlsx(content: bytes) -> None:
    if not content or len(content) > MAX_COMPRESSED_BYTES:
        raise _blocked("INVALID_XLSX_PACKAGE", "compressed workbook size is outside the limit")
    try:
        archive = ZipFile(BytesIO(content))
    except (BadZipFile, OSError, ValueError) as exc:
        raise _blocked("INVALID_XLSX_PACKAGE", "input is not a valid xlsx ZIP package") from exc

    with archive:
        entries = archive.infolist()
        if not entries or len(entries) > MAX_PACKAGE_ENTRIES:
            raise _blocked("XLSX_PACKAGE_LIMIT", "package entry count is outside the limit")
        names = [entry.filename for entry in entries]
        if len(names) != len(set(names)) or len(names) != len({name.casefold() for name in names}):
            raise _blocked("DUPLICATE_PACKAGE_ENTRY", "duplicate ZIP entry names are forbidden")
        required = {"[Content_Types].xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
        if not required.issubset(names):
            raise _blocked("INVALID_XLSX_PACKAGE", "required workbook parts are missing")

        total_uncompressed = 0
        for entry in entries:
            if _unsafe_package_path(entry.filename):
                raise _blocked("UNSAFE_PACKAGE_PATH", "unsafe ZIP entry path is forbidden")
            if entry.flag_bits & 0x1:
                raise _blocked("ENCRYPTED_XLSX_BLOCKED", "encrypted workbook entries are forbidden")
            if _is_active_content(entry.filename):
                raise _blocked("ACTIVE_CONTENT_BLOCKED", "active workbook content is forbidden")
            if entry.file_size > MAX_ENTRY_BYTES:
                raise _blocked("XLSX_PACKAGE_LIMIT", "one package entry exceeds the size limit")
            total_uncompressed += entry.file_size
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise _blocked("XLSX_PACKAGE_LIMIT", "uncompressed workbook exceeds the size limit")
            if (
                entry.file_size
                and entry.file_size / max(entry.compress_size, 1) > MAX_COMPRESSION_RATIO
            ):
                raise _blocked("XLSX_PACKAGE_LIMIT", "package compression ratio exceeds the limit")

        for entry in entries:
            lowered = entry.filename.casefold()
            if not (lowered.endswith(".xml") or lowered.endswith(".rels")):
                continue
            payload = archive.read(entry)
            payload_lower = payload.lower()
            if b"<!doctype" in payload_lower or b"<!entity" in payload_lower:
                raise _blocked("XML_ENTITY_BLOCKED", "DTD and entity declarations are forbidden")
            if lowered.endswith(".rels"):
                try:
                    relationship_root = ET.fromstring(payload)
                except ET.ParseError as exc:
                    raise _blocked("INVALID_XLSX_PACKAGE", "relationship XML is malformed") from exc
                for relationship in relationship_root:
                    if relationship.attrib.get("TargetMode", "").casefold() == "external":
                        raise _blocked(
                            "EXTERNAL_RELATIONSHIP_BLOCKED",
                            "external workbook relationships are forbidden",
                        )
                    relationship_type = relationship.attrib.get("Type", "")
                    if _declares_active_content(relationship_type):
                        raise _blocked(
                            "ACTIVE_CONTENT_BLOCKED",
                            "active workbook relationship is forbidden",
                        )
            if entry.filename == "[Content_Types].xml":
                try:
                    content_types_root = ET.fromstring(payload)
                except ET.ParseError as exc:
                    raise _blocked(
                        "INVALID_XLSX_PACKAGE", "content-types XML is malformed"
                    ) from exc
                for declaration in content_types_root:
                    if _declares_active_content(declaration.attrib.get("ContentType", "")):
                        raise _blocked(
                            "ACTIVE_CONTENT_BLOCKED",
                            "active workbook content type is forbidden",
                        )


def _load_rows(
    content: bytes,
    profile: WorkbookProfile,
) -> tuple[tuple[RawSourceRow, ...], tuple[tuple[object, ...], ...]]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise _blocked("INVALID_XLSX_PACKAGE", "workbook could not be loaded safely") from exc

    try:
        if tuple(workbook.sheetnames) != (profile.sheet_name,):
            raise _blocked("SHEET_MISMATCH", "workbook sheet list does not match the profile")
        worksheet = workbook[profile.sheet_name]
        if worksheet.sheet_state != "visible":
            raise _blocked("SHEET_MISMATCH", "profile sheet must be visible")
        if worksheet.max_row > MAX_WORKSHEET_ROWS or worksheet.max_column > MAX_WORKSHEET_COLUMNS:
            raise _blocked("WORKSHEET_LIMIT", "worksheet dimensions exceed parser limits")
        if worksheet.max_column != len(profile.headers):
            raise _blocked("HEADER_MISMATCH", "worksheet column count does not match the profile")

        raw_rows: list[RawSourceRow] = []
        values_by_row: list[tuple[object, ...]] = []
        header: tuple[object, ...] | None = None
        for row_number, cells in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
            ),
            start=1,
        ):
            for cell in cells:
                if cell.data_type == "f":
                    raise _blocked(
                        "FORMULA_BLOCKED",
                        "formula cells are forbidden in source workbooks",
                        row=cell.row,
                        column=str(cell.column_letter),
                    )
                if cell.data_type == "e":
                    raise _blocked(
                        "CELL_ERROR_BLOCKED",
                        "error cells are forbidden in source workbooks",
                        row=cell.row,
                        column=str(cell.column_letter),
                    )
            values = tuple(cell.value for cell in cells)
            if row_number == 1:
                header = values
                continue
            if all(value is None or value == "" for value in values):
                continue
            raw_rows.append(
                RawSourceRow(
                    sheet_ref=profile.sheet_name,
                    source_row_number=row_number,
                    raw_values=tuple(_raw_value(value) for value in values),
                )
            )
            values_by_row.append(values)

        if header != profile.headers:
            raise _blocked("HEADER_MISMATCH", "worksheet header does not exactly match the profile")
        if not raw_rows:
            raise _blocked("EMPTY_WORKBOOK", "workbook contains no source rows")
        return tuple(raw_rows), tuple(values_by_row)
    finally:
        workbook.close()


def _raw_value(value: object) -> RawCellValue:
    if value is None or isinstance(value, (str, int, float, bool, date, datetime, time)):
        return value
    return str(value)


def _require_text(value: object, *, row: int, column: str, optional: bool = False) -> str | None:
    if value is None or value == "":
        if optional:
            return None
        raise _blocked(
            "REQUIRED_CELL_MISSING", "required text cell is blank", row=row, column=column
        )
    if not isinstance(value, str):
        raise _blocked("CELL_TYPE_MISMATCH", "expected a text cell", row=row, column=column)
    if value != value.strip() or not value:
        raise _blocked(
            "CELL_VALUE_INVALID",
            "text must be nonblank without edge whitespace",
            row=row,
            column=column,
        )
    return value


def _require_date(value: object, *, row: int, column: str) -> date:
    if isinstance(value, datetime):
        if value.time() != time.min:
            raise _blocked(
                "CELL_VALUE_INVALID",
                "date cell contains a non-midnight time",
                row=row,
                column=column,
            )
        return value.date()
    if isinstance(value, date):
        return value
    raise _blocked("CELL_TYPE_MISMATCH", "expected an Excel date cell", row=row, column=column)


def _require_hh_mm(value: object, *, row: int, column: str) -> time:
    text_value = _require_text(value, row=row, column=column)
    assert text_value is not None
    if not _HH_MM.fullmatch(text_value):
        raise _blocked("CELL_VALUE_INVALID", "expected exact HH:MM text", row=row, column=column)
    return datetime.strptime(text_value, "%H:%M").time()


def _require_integer(
    value: object,
    *,
    row: int,
    column: str,
    text_only: bool = False,
) -> int:
    if text_only:
        text_value = _require_text(value, row=row, column=column)
        assert text_value is not None
        if not re.fullmatch(r"\d+", text_value):
            raise _blocked(
                "CELL_VALUE_INVALID", "expected unsigned integer text", row=row, column=column
            )
        return int(text_value)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise _blocked(
            "CELL_TYPE_MISMATCH", "expected a numeric integer cell", row=row, column=column
        )
    if not math.isfinite(float(value)) or not float(value).is_integer():
        raise _blocked(
            "CELL_VALUE_INVALID", "numeric cell must be a finite integer", row=row, column=column
        )
    return int(value)


def _signature(values: tuple[object, ...]) -> str:
    canonical = json.dumps(values, ensure_ascii=False, separators=(",", ":"), default=str)
    return sha256(canonical.encode("utf-8")).hexdigest()


def _certification_number(value: str, *, row: int, column: str) -> str:
    try:
        return normalize_certification_number(value)
    except ValueError as exc:
        raise _blocked(
            "CELL_VALUE_INVALID",
            "certification number is outside the canonical input family",
            row=row,
            column=column,
        ) from exc


def _date_time_minutes(start: time, end: time, *, row: int) -> int:
    start_seconds = start.hour * 3600 + start.minute * 60
    end_seconds = end.hour * 3600 + end.minute * 60
    if end_seconds <= start_seconds:
        raise _blocked("TIME_RANGE_INVALID", "planned end must be after planned start", row=row)
    minutes = (end_seconds - start_seconds) // 60
    if minutes % 30:
        raise _blocked(
            "TIME_RANGE_INVALID", "planned duration must use the 30-minute grid", row=row
        )
    return minutes


def parse_nhis_schedule_workbook(
    content: bytes,
    *,
    target_month: date,
    original_filename: str | None = None,
) -> ParsedWorkbook[NhisScheduleRow]:
    """Parse one approved monthly NHIS schedule workbook into an immutable preview."""

    del original_filename
    if target_month.day != 1:
        raise _blocked("TARGET_MONTH_INVALID", "target_month must be the first day of a month")
    raw_rows, values_by_row = _load_rows(content, NHIS_SCHEDULE_PROFILE_V1)
    parsed: list[NhisScheduleRow] = []
    signature_counts: Counter[str] = Counter()

    for raw_row, values in zip(raw_rows, values_by_row, strict=True):
        row = raw_row.source_row_number
        service_date = _require_date(values[0], row=row, column="일자")
        if (service_date.year, service_date.month) != (target_month.year, target_month.month):
            raise _blocked(
                "TARGET_MONTH_MISMATCH",
                "source row date is outside the selected month",
                row=row,
                column="일자",
            )
        planned_start = _require_hh_mm(values[1], row=row, column="시작시간")
        planned_end = _require_hh_mm(values[2], row=row, column="종료시간")
        declared_minutes = _date_time_minutes(planned_start, planned_end, row=row)
        recipient_name = _require_text(values[3], row=row, column="수급자명")
        recipient_number_raw = _require_text(values[4], row=row, column="수급자\n인정번호")
        staff_name = _require_text(values[5], row=row, column="요양보호사명")
        staff_birth_date = _require_date(values[6], row=row, column="생년월일")
        staff_number = _require_text(values[7], row=row, column="요양보호사번호")
        worker_category = _require_text(values[8], row=row, column="종사자구분")
        family_flag = _require_text(values[9], row=row, column="가족여부")
        family_relationship = _require_text(
            values[10],
            row=row,
            column="가족관계",
            optional=True,
        )
        service_category = _require_text(values[11], row=row, column="서비스구분")
        fee_code = _require_text(values[12], row=row, column="수가코드")
        fee_name = _require_text(values[13], row=row, column="수가명")
        fee_amount = _require_integer(values[14], row=row, column="수가")
        assert recipient_name is not None
        assert recipient_number_raw is not None
        assert staff_name is not None
        assert staff_number is not None
        assert worker_category is not None
        assert family_flag is not None
        assert service_category is not None
        assert fee_code is not None
        assert fee_name is not None
        recipient_number = _certification_number(
            recipient_number_raw,
            row=row,
            column="수급자\n인정번호",
        )
        if family_flag not in {"Y", "N"}:
            raise _blocked(
                "CELL_VALUE_INVALID", "family flag must be Y or N", row=row, column="가족여부"
            )
        if (family_flag == "Y") != (family_relationship is not None):
            raise _blocked(
                "CELL_VALUE_INVALID",
                "family relationship presence must agree with family flag",
                row=row,
                column="가족관계",
            )
        if service_category not in VISIT_SERVICE_TYPES:
            raise _blocked(
                "CELL_VALUE_INVALID",
                "service category is outside the approved profile",
                row=row,
                column="서비스구분",
            )
        if fee_amount < 0:
            raise _blocked(
                "CELL_VALUE_INVALID", "fee amount cannot be negative", row=row, column="수가"
            )

        signature = _signature(
            (
                service_date.isoformat(),
                planned_start.isoformat(timespec="minutes"),
                planned_end.isoformat(timespec="minutes"),
                recipient_name,
                recipient_number,
                staff_name,
                staff_birth_date.isoformat(),
                staff_number,
                worker_category,
                family_flag,
                family_relationship,
                service_category,
                fee_code,
                fee_name,
                fee_amount,
            )
        )
        signature_counts[signature] += 1
        parsed.append(
            NhisScheduleRow(
                sheet_ref=NHIS_SCHEDULE_PROFILE_V1.sheet_name,
                source_row_number=row,
                service_date=service_date,
                planned_start=planned_start,
                planned_end=planned_end,
                declared_minutes=declared_minutes,
                recipient_name=recipient_name,
                recipient_certification_number=recipient_number,
                staff_name=staff_name,
                staff_birth_date=staff_birth_date,
                staff_external_number=staff_number,
                worker_category=worker_category,
                family_flag=family_flag,
                family_relationship=family_relationship,
                service_category=service_category,
                fee_code=fee_code,
                fee_name=fee_name,
                fee_amount=fee_amount,
                occurrence_signature=signature,
                occurrence_ordinal=signature_counts[signature],
            )
        )

    immutable_rows = tuple(parsed)
    return ParsedWorkbook(
        profile_version=NHIS_SCHEDULE_PROFILE_V1.profile_version,
        source_type=NHIS_SCHEDULE_PROFILE_V1.source_type,
        content_digest=sha256(content).hexdigest(),
        sheet_ref=NHIS_SCHEDULE_PROFILE_V1.sheet_name,
        raw_rows=raw_rows,
        parsed_rows=immutable_rows,
        target_rows=immutable_rows,
    )


def _require_timestamp(value: object, *, row: int, column: str) -> datetime:
    text_value = _require_text(value, row=row, column=column)
    assert text_value is not None
    if not _TIMESTAMP.fullmatch(text_value):
        raise _blocked(
            "CELL_VALUE_INVALID",
            "expected exact YYYY-MM-DD HH:MM:SS text",
            row=row,
            column=column,
        )
    try:
        parsed = datetime.strptime(text_value, "%Y-%m-%d %H:%M:%S")
    except ValueError as exc:
        raise _blocked(
            "CELL_VALUE_INVALID", "timestamp is not a calendar value", row=row, column=column
        ) from exc
    return parsed.replace(tzinfo=KST)


def parse_rfid_workbook(
    content: bytes,
    *,
    target_date: date,
    original_filename: str | None = None,
) -> ParsedWorkbook[RfidRow]:
    """Parse a range export while deriving exactly one selected-day RFID snapshot."""

    del original_filename
    raw_rows, values_by_row = _load_rows(content, RFID_PROFILE_V1)
    parsed: list[RfidRow] = []
    signature_counts: Counter[str] = Counter()

    for raw_row, values in zip(raw_rows, values_by_row, strict=True):
        row = raw_row.source_row_number
        transmission_kind = _require_text(values[0], row=row, column="구분")
        recipient_name = _require_text(values[1], row=row, column="수급자성명")
        recipient_number_raw = _require_text(values[2], row=row, column="인정번호")
        staff_name = _require_text(values[3], row=row, column="요양요원")
        staff_phone = _require_text(values[4], row=row, column="핸드폰번호")
        service_category = _require_text(values[5], row=row, column="급여종류")
        reference_minutes = _require_integer(values[6], row=row, column="총시간", text_only=True)
        actual_start = _require_timestamp(values[7], row=row, column="시작시간")
        end_text = _require_text(values[8], row=row, column="종료시간", optional=True)
        actual_end = (
            _require_timestamp(end_text, row=row, column="종료시간")
            if end_text is not None
            else None
        )
        use_state = _require_text(values[9], row=row, column="사용여부")
        assert transmission_kind is not None
        assert recipient_name is not None
        assert recipient_number_raw is not None
        assert staff_name is not None
        assert staff_phone is not None
        assert service_category is not None
        assert use_state is not None
        recipient_number = _certification_number(
            recipient_number_raw,
            row=row,
            column="인정번호",
        )
        try:
            canonical_phone, normalized_phone = normalize_phone_number(staff_phone)
        except ValueError as exc:
            raise _blocked(
                "CELL_VALUE_INVALID",
                "staff phone is outside the canonical input family",
                row=row,
                column="핸드폰번호",
            ) from exc
        assert canonical_phone is not None
        assert normalized_phone is not None
        if service_category not in VISIT_SERVICE_TYPES:
            raise _blocked(
                "CELL_VALUE_INVALID",
                "service category is outside the approved profile",
                row=row,
                column="급여종류",
            )
        if transmission_kind == "시작전송":
            if actual_end is not None or use_state != "미사용" or reference_minutes != 0:
                raise _blocked(
                    "RFID_STATE_INVALID",
                    "start-only rows must have blank end, zero reference minutes, and unused state",
                    row=row,
                )
            event_state = RfidEventState.START_ONLY
        elif transmission_kind == "자동전송":
            if actual_end is None or use_state != "사용":
                raise _blocked(
                    "RFID_STATE_INVALID",
                    "automatic rows must have an end timestamp and used state",
                    row=row,
                )
            if actual_end <= actual_start:
                raise _blocked(
                    "TIME_RANGE_INVALID",
                    "RFID end must be after start",
                    row=row,
                    column="종료시간",
                )
            event_state = RfidEventState.COMPLETE
        else:
            raise _blocked(
                "CELL_VALUE_INVALID",
                "transmission kind is outside the approved profile",
                row=row,
                column="구분",
            )

        signature = _signature(
            (
                transmission_kind,
                recipient_name,
                recipient_number,
                staff_name,
                canonical_phone,
                service_category,
                reference_minutes,
                actual_start.isoformat(),
                actual_end.isoformat() if actual_end is not None else None,
                use_state,
            )
        )
        signature_counts[signature] += 1
        parsed.append(
            RfidRow(
                sheet_ref=RFID_PROFILE_V1.sheet_name,
                source_row_number=row,
                transmission_kind=transmission_kind,
                recipient_name=recipient_name,
                recipient_certification_number=recipient_number,
                staff_name=staff_name,
                staff_phone=canonical_phone,
                staff_phone_normalized=normalized_phone,
                service_category=service_category,
                reference_minutes=reference_minutes,
                actual_start=actual_start,
                actual_end=actual_end,
                use_state=use_state,
                event_state=event_state,
                occurrence_signature=signature,
                occurrence_ordinal=signature_counts[signature],
            )
        )

    immutable_rows = tuple(parsed)
    target_rows = tuple(row for row in immutable_rows if row.actual_start.date() == target_date)
    if not target_rows:
        raise _blocked("TARGET_DATE_NOT_FOUND", "workbook contains no rows for the selected day")
    warnings = ("EXPORT_CONTAINS_OTHER_DATES",) if len(target_rows) != len(immutable_rows) else ()
    return ParsedWorkbook(
        profile_version=RFID_PROFILE_V1.profile_version,
        source_type=RFID_PROFILE_V1.source_type,
        content_digest=sha256(content).hexdigest(),
        sheet_ref=RFID_PROFILE_V1.sheet_name,
        raw_rows=raw_rows,
        parsed_rows=immutable_rows,
        target_rows=target_rows,
        warning_codes=warnings,
    )
