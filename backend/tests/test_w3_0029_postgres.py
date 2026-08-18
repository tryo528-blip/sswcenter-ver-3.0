"""Live PostgreSQL and HTTP proofs for the W3 0029 workspace.

The module is inert unless the dedicated disposable harness exports
``SSWCENTER_W3_0029_REAL_PG=1``.  It uses only the approved pseudonymous
workbook fixture and never rewrites that fixture on disk.
"""

from __future__ import annotations

import json
import os
import time
from collections.abc import Iterator
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from threading import Barrier, Event
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import Engine, create_engine, text
from sqlalchemy.engine import make_url
from sqlalchemy.exc import IntegrityError, ProgrammingError
from sqlalchemy.orm import Session, sessionmaker

from app.api.dependencies import get_current_account, get_db_session, require_csrf
from app.core.auth import CurrentAccount
from app.core.settings import Settings, assert_safe_test_database_url, get_settings
from app.db.postcheck_current_0029 import (
    CURRENT_0029_MARKER,
    EXPECTED_REVISION,
    HEAD_MARKER,
    verify_current_0029,
)
from app.domains.recipient.errors import RecipientDomainError
from app.domains.w3.matching_repository import W3MatchingRepository
from app.domains.w3.schemas import (
    W3ApplyRequest,
    W3ConfirmRequest,
    W3PlanAdjustmentRequest,
    W3PlanAdjustmentResponse,
    W3ResolveDecisionRequest,
    W3SourceType,
    W3SupplementAction,
    W3SupplementRequest,
    W3SupplementResponse,
)
from app.domains.w3.service import XLSX_MEDIA_TYPE, W3Service
from app.main import create_app

pytestmark = pytest.mark.skipif(
    os.getenv("SSWCENTER_W3_0029_REAL_PG") != "1",
    reason="requires the dedicated W3 0029 disposable PostgreSQL harness",
)

TARGET_DATE = date(2026, 7, 6)
SCHEDULE_MONTH = date(2026, 7, 1)
START_ONLY_SOURCE_ROW = 174
START_ONLY_CERTIFICATION_NUMBER = "L9586300259"
FIXTURE = Path(__file__).parent / "fixtures" / "w3" / "workbooks" / "rfid_202607_v1.xlsx"


def _required_url(name: str) -> str:
    value = os.getenv(name)
    assert value, f"{name} must be explicitly exported"
    assert_safe_test_database_url(value)
    return value


def _sqlstate(error: BaseException) -> str | None:
    original = getattr(error, "orig", None)
    return None if original is None else str(getattr(original, "sqlstate", None) or "")


@pytest.fixture(scope="module")
def owner_engine() -> Iterator[Engine]:
    engine = create_engine(_required_url("SSWCENTER_DATABASE_URL"), pool_pre_ping=True)
    try:
        yield engine
    finally:
        engine.dispose()


@pytest.fixture(scope="module")
def app_engine() -> Iterator[Engine]:
    engine = create_engine(_required_url("SSWCENTER_APP_DATABASE_URL"), pool_pre_ping=True)
    try:
        yield engine
    finally:
        engine.dispose()


@pytest.fixture(scope="module")
def superuser_engine() -> Iterator[Engine]:
    url = make_url(_required_url("SSWCENTER_DATABASE_URL")).set(username="postgres")
    engine = create_engine(url, pool_pre_ping=True)
    try:
        yield engine
    finally:
        engine.dispose()


@pytest.fixture(scope="module")
def settings() -> Settings:
    data_root = os.getenv("SSWCENTER_DATA_ROOT")
    assert data_root, "SSWCENTER_DATA_ROOT must be explicitly exported"
    return Settings(
        environment="test",
        database_url=_required_url("SSWCENTER_APP_DATABASE_URL"),
        data_root=Path(data_root),
    )


@dataclass(frozen=True, slots=True)
class SeededWorker:
    staff_id: int
    employment_id: int
    care_assignment_id: int
    w2_schedule_id: int


@dataclass(frozen=True, slots=True)
class SeededCase:
    account: CurrentAccount
    recipient_id: int
    certification_period_id: int
    service_type_id: int
    recipient_contract_id: int
    workers: tuple[SeededWorker, SeededWorker]
    wrong_date_schedule_id: int


@pytest.fixture(scope="module")
def seeded(owner_engine: Engine) -> SeededCase:
    suffix = uuid4().hex
    with owner_engine.begin() as connection:
        def insert_staff(label: str) -> int:
            value = connection.scalar(
                text(
                    """
                    INSERT INTO erp.staff (name, display_name, birth_date, sex_code)
                    VALUES (:name, :name, DATE '1990-01-01', 'TEST')
                    RETURNING id
                    """
                ),
                {"name": f"W3 0029 {label} {suffix}"},
            )
            assert value is not None
            return int(value)

        actor_staff_id = insert_staff("actor")
        worker_staff_ids = (insert_staff("worker A"), insert_staff("worker B"))
        account_id = connection.scalar(
            text(
                """
                INSERT INTO erp.user_account (
                    staff_id, account_code, display_name, role_code,
                    pin_hash, pin_lookup_hmac, pin_key_version
                ) VALUES (
                    :staff_id, :account_code, :display_name, 'ADMIN',
                    :pin_hash, :pin_lookup_hmac, 1
                ) RETURNING id
                """
            ),
            {
                "staff_id": actor_staff_id,
                "account_code": f"W3-0029-{suffix}",
                "display_name": f"W3 0029 actor {suffix}",
                "pin_hash": f"unused-{suffix}",
                "pin_lookup_hmac": sha256(suffix.encode("ascii")).digest(),
            },
        )
        assert account_id is not None
        account_id = int(account_id)

        first_sequence = int(
            connection.scalar(
                text("SELECT COALESCE(MAX(staff_no_sequence), 0) + 1 FROM erp.staff_employment")
            )
            or 1
        )
        worker_employment_ids: list[int] = []
        for offset, staff_id in enumerate(worker_staff_ids):
            employment_id = connection.scalar(
                text(
                    """
                    INSERT INTO erp.staff_employment (
                        staff_id, employment_no, staff_no, staff_no_year,
                        staff_no_sequence, start_date, end_date,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :staff_id, 1, :staff_no, 2026, :staff_no_sequence,
                        DATE '2026-01-01', DATE '2026-12-31',
                        :account_id, :account_id
                    ) RETURNING id
                    """
                ),
                {
                    "staff_id": staff_id,
                    "staff_no": f"W3-{offset + 1}-{suffix}",
                    "staff_no_sequence": first_sequence + offset,
                    "account_id": account_id,
                },
            )
            assert employment_id is not None
            worker_employment_ids.append(int(employment_id))
            connection.execute(
                text(
                    """
                    INSERT INTO erp.staff_position_period (
                        staff_id, employment_id, position_code, start_date, end_date,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :staff_id, :employment_id, 'CARE_WORKER',
                        DATE '2026-01-01', DATE '2026-12-31',
                        :account_id, :account_id
                    )
                    """
                ),
                {
                    "staff_id": staff_id,
                    "employment_id": employment_id,
                    "account_id": account_id,
                },
            )

        service_type_id = connection.scalar(
            text("SELECT id FROM erp.service_type WHERE code = 'HOME_CARE' AND active")
        )
        assert service_type_id is not None
        service_type_id = int(service_type_id)
        for staff_id, employment_id in zip(
            worker_staff_ids, worker_employment_ids, strict=True
        ):
            connection.execute(
                text(
                    """
                    INSERT INTO erp.staff_service_qualification_period (
                        staff_id, employment_id, service_type_id,
                        start_date, end_date, source_license_id,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :staff_id, :employment_id, :service_type_id,
                        DATE '2026-01-01', DATE '2026-12-31', NULL,
                        :account_id, :account_id
                    )
                    """
                ),
                {
                    "staff_id": staff_id,
                    "employment_id": employment_id,
                    "service_type_id": service_type_id,
                    "account_id": account_id,
                },
            )

        recipient_id = connection.scalar(
            text(
                """
                INSERT INTO erp.recipient (
                    name, birth_date, sex_code, mobile_phone,
                    created_by_account_id, updated_by_account_id
                ) VALUES (
                    :name, DATE '1950-01-01', 'TEST', :mobile,
                    :account_id, :account_id
                ) RETURNING id
                """
            ),
            {
                "name": f"W3 0029 recipient {suffix}",
                "mobile": f"010{int(suffix[:8], 16) % 100_000_000:08d}",
                "account_id": account_id,
            },
        )
        assert recipient_id is not None
        recipient_id = int(recipient_id)
        connection.execute(
            text(
                """
                INSERT INTO erp.recipient_certification_identity (
                    recipient_id, certification_number,
                    created_by_account_id, updated_by_account_id
                ) VALUES (:recipient_id, :number, :account_id, :account_id)
                """
            ),
            {
                "recipient_id": recipient_id,
                "number": START_ONLY_CERTIFICATION_NUMBER,
                "account_id": account_id,
            },
        )
        certification_period_id = connection.scalar(
            text(
                """
                INSERT INTO erp.recipient_certification_period (
                    recipient_id, grade_code, start_date, end_date,
                    created_by_account_id, updated_by_account_id
                ) VALUES (
                    :recipient_id, '3', DATE '2026-01-01', DATE '2026-12-31',
                    :account_id, :account_id
                ) RETURNING id
                """
            ),
            {"recipient_id": recipient_id, "account_id": account_id},
        )
        assert certification_period_id is not None
        certification_period_id = int(certification_period_id)
        recipient_contract_id = connection.scalar(
            text(
                """
                INSERT INTO erp.recipient_contract (
                    recipient_id, service_type_id, start_date, end_date,
                    created_by_account_id, updated_by_account_id
                ) VALUES (
                    :recipient_id, :service_type_id,
                    DATE '2026-01-01', DATE '2026-12-31',
                    :account_id, :account_id
                ) RETURNING id
                """
            ),
            {
                "recipient_id": recipient_id,
                "service_type_id": service_type_id,
                "account_id": account_id,
            },
        )
        assert recipient_contract_id is not None
        recipient_contract_id = int(recipient_contract_id)

        care_assignment_ids: list[int] = []
        for staff_id, employment_id in zip(
            worker_staff_ids, worker_employment_ids, strict=True
        ):
            assignment_id = connection.scalar(
                text(
                    """
                    INSERT INTO erp.care_assignment (
                        recipient_contract_id, staff_id, employment_id,
                        assignment_kind, family_relationship_text,
                        start_date, end_date,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :contract_id, :staff_id, :employment_id,
                        'GENERAL', NULL, DATE '2026-01-01', DATE '2026-12-31',
                        :account_id, :account_id
                    ) RETURNING id
                    """
                ),
                {
                    "contract_id": recipient_contract_id,
                    "staff_id": staff_id,
                    "employment_id": employment_id,
                    "account_id": account_id,
                },
            )
            assert assignment_id is not None
            care_assignment_ids.append(int(assignment_id))

        connection.execute(
            text(
                """
                INSERT INTO erp.w2_schedule_month_control (
                    schedule_month, created_by_account_id, updated_by_account_id
                ) VALUES (:month, :account_id, :account_id)
                """
            ),
            {"month": SCHEDULE_MONTH, "account_id": account_id},
        )
        schedule_ids: list[int] = []
        schedule_windows = (
            (
                datetime(2026, 7, 6, 2, 30, tzinfo=UTC),
                datetime(2026, 7, 6, 3, 30, tzinfo=UTC),
            ),
            (
                datetime(2026, 7, 6, 4, 0, tzinfo=UTC),
                datetime(2026, 7, 6, 5, 0, tzinfo=UTC),
            ),
        )
        for staff_id, employment_id, (starts_at_utc, ends_at_utc) in zip(
            worker_staff_ids,
            worker_employment_ids,
            schedule_windows,
            strict=True,
        ):
            schedule_id = connection.scalar(
                text(
                    """
                    INSERT INTO erp.w2_schedule (
                        schedule_month, recipient_id, service_type_id,
                        starts_at_utc, ends_at_utc,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :month, :recipient_id, :service_type_id,
                        :starts_at_utc, :ends_at_utc,
                        :account_id, :account_id
                    ) RETURNING id
                    """
                ),
                {
                    "month": SCHEDULE_MONTH,
                    "recipient_id": recipient_id,
                    "service_type_id": service_type_id,
                    "starts_at_utc": starts_at_utc,
                    "ends_at_utc": ends_at_utc,
                    "account_id": account_id,
                },
            )
            assert schedule_id is not None
            schedule_id = int(schedule_id)
            schedule_ids.append(schedule_id)
            connection.execute(
                text(
                    """
                    INSERT INTO erp.w2_schedule_staff (
                        schedule_id, staff_id, employment_id,
                        created_by_account_id, updated_by_account_id
                    ) VALUES (
                        :schedule_id, :staff_id, :employment_id,
                        :account_id, :account_id
                    )
                    """
                ),
                {
                    "schedule_id": schedule_id,
                    "staff_id": staff_id,
                    "employment_id": employment_id,
                    "account_id": account_id,
                },
            )

        wrong_date_schedule_id = connection.scalar(
            text(
                """
                INSERT INTO erp.w2_schedule (
                    schedule_month, recipient_id, service_type_id,
                    starts_at_utc, ends_at_utc,
                    created_by_account_id, updated_by_account_id
                ) VALUES (
                    :month, :recipient_id, :service_type_id,
                    :starts_at_utc, :ends_at_utc,
                    :account_id, :account_id
                ) RETURNING id
                """
            ),
            {
                "month": SCHEDULE_MONTH,
                "recipient_id": recipient_id,
                "service_type_id": service_type_id,
                "starts_at_utc": datetime(2026, 7, 7, 2, 30, tzinfo=UTC),
                "ends_at_utc": datetime(2026, 7, 7, 3, 30, tzinfo=UTC),
                "account_id": account_id,
            },
        )
        assert wrong_date_schedule_id is not None
        wrong_date_schedule_id = int(wrong_date_schedule_id)
        connection.execute(
            text(
                """
                INSERT INTO erp.w2_schedule_staff (
                    schedule_id, staff_id, employment_id,
                    created_by_account_id, updated_by_account_id
                ) VALUES (
                    :schedule_id, :staff_id, :employment_id,
                    :account_id, :account_id
                )
                """
            ),
            {
                "schedule_id": wrong_date_schedule_id,
                "staff_id": worker_staff_ids[0],
                "employment_id": worker_employment_ids[0],
                "account_id": account_id,
            },
        )

    workers = tuple(
        SeededWorker(staff_id, employment_id, assignment_id, schedule_id)
        for staff_id, employment_id, assignment_id, schedule_id in zip(
            worker_staff_ids,
            worker_employment_ids,
            care_assignment_ids,
            schedule_ids,
            strict=True,
        )
    )
    assert len(workers) == 2
    return SeededCase(
        account=CurrentAccount(
            id=account_id,
            display_name=f"W3 0029 actor {suffix}",
            role_code="ADMIN",
        ),
        recipient_id=recipient_id,
        certification_period_id=certification_period_id,
        service_type_id=service_type_id,
        recipient_contract_id=recipient_contract_id,
        workers=(workers[0], workers[1]),
        wrong_date_schedule_id=wrong_date_schedule_id,
    )


def _factory(engine: Engine) -> sessionmaker[Session]:
    return sessionmaker(bind=engine, expire_on_commit=False)


def _single_start_only_workbook(*, title: str) -> bytes:
    workbook = load_workbook(BytesIO(FIXTURE.read_bytes()))
    worksheet = workbook[workbook.sheetnames[0]]
    if worksheet.max_row > START_ONLY_SOURCE_ROW:
        worksheet.delete_rows(
            START_ONLY_SOURCE_ROW + 1,
            worksheet.max_row - START_ONLY_SOURCE_ROW,
        )
    worksheet.delete_rows(2, START_ONLY_SOURCE_ROW - 2)
    workbook.properties.title = title
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _domain_code(
    call: object,
    expected_code: str,
    expected_status: int,
) -> RecipientDomainError:
    assert callable(call)
    with pytest.raises(RecipientDomainError) as caught:
        call()
    assert caught.value.code == expected_code
    assert caught.value.status_code == expected_status
    return caught.value


@dataclass(frozen=True, slots=True)
class PreparedRun:
    run_id: int
    apply_request: W3ApplyRequest


def _prepare_confirmed_run(
    *,
    factory: sessionmaker[Session],
    settings: Settings,
    seeded: SeededCase,
    worker: SeededWorker,
    content: bytes,
    label: str,
) -> PreparedRun:
    with factory() as session:
        uploaded = W3Service(session, settings).upload_workbook(
            content=content,
            original_filename=f"w3-{label}.xlsx",
            source_type=W3SourceType.RFID,
            target_date=TARGET_DATE,
            account=seeded.account,
        )
    run = uploaded.latest_run
    assert run is not None
    assert run.status.value == "PREVIEW_READY"
    assert run.counts.raw_rows == run.counts.normalized_rows == run.counts.target_rows == 1
    assert run.counts.review_pending == 1
    assert len(run.decisions) == 1
    decision = run.decisions[0]

    invalid_request = W3ResolveDecisionRequest(
        expected_run_row_version=run.row_version,
        command_idempotency_key=f"w3-{label}-invalid-link",
        recipient_id=9_999_999_999,
        certification_period_id=seeded.certification_period_id,
        staff_id=worker.staff_id,
        employment_id=worker.employment_id,
        service_type_id=seeded.service_type_id,
        recipient_contract_id=seeded.recipient_contract_id,
        care_assignment_id=worker.care_assignment_id,
        w2_schedule_id=worker.w2_schedule_id,
    )
    with factory() as session:
        service = W3Service(session, settings)
        _domain_code(
            lambda: service.resolve_decision(
                run.id, decision.id, invalid_request, seeded.account
            ),
            "W3_TYPED_LINK_INVALID",
            422,
        )

    resolve_request = invalid_request.model_copy(
        update={
            "command_idempotency_key": f"w3-{label}-resolve-fixed",
            "recipient_id": seeded.recipient_id,
        }
    )
    with factory() as session:
        resolved = W3Service(session, settings).resolve_decision(
            run.id, decision.id, resolve_request, seeded.account
        )
    resolved_run = resolved.latest_run
    assert resolved_run is not None and resolved_run.id == run.id
    assert resolved_run.counts.review_pending == 0
    assert resolved_run.counts.manual_matches == 1
    assert resolved_run.row_version == run.row_version + 1
    assert resolved_run.can_confirm

    with factory() as session:
        retry = W3Service(session, settings).resolve_decision(
            run.id, decision.id, resolve_request, seeded.account
        )
    assert retry.latest_run is not None
    assert retry.latest_run.row_version == resolved_run.row_version
    with factory() as session:
        service = W3Service(session, settings)
        _domain_code(
            lambda: service.resolve_decision(
                run.id,
                decision.id,
                resolve_request.model_copy(
                    update={"expected_run_row_version": resolved_run.row_version}
                ),
                seeded.account,
            ),
            "W3_IDEMPOTENCY_CONFLICT",
            409,
        )

    assert resolved_run.preview_digest is not None
    confirm_request = W3ConfirmRequest(
        expected_row_version=resolved_run.row_version,
        preview_digest=resolved_run.preview_digest,
        command_idempotency_key=f"w3-{label}-confirm-fixed",
    )
    with factory() as session:
        confirmed = W3Service(session, settings).confirm_run(
            run.id, confirm_request, seeded.account
        )
    confirmed_run = confirmed.latest_run
    assert confirmed_run is not None and confirmed_run.id == run.id
    assert confirmed_run.status.value == "CONFIRMED"
    assert confirmed_run.row_version == resolved_run.row_version + 1
    assert confirmed_run.can_apply

    with factory() as session:
        retry = W3Service(session, settings).confirm_run(
            run.id, confirm_request, seeded.account
        )
    assert retry.latest_run is not None
    assert retry.latest_run.row_version == confirmed_run.row_version
    with factory() as session:
        service = W3Service(session, settings)
        _domain_code(
            lambda: service.confirm_run(
                run.id,
                confirm_request.model_copy(
                    update={"expected_row_version": confirmed_run.row_version}
                ),
                seeded.account,
            ),
            "W3_IDEMPOTENCY_CONFLICT",
            409,
        )
    return PreparedRun(
        run_id=run.id,
        apply_request=W3ApplyRequest(
            expected_row_version=confirmed_run.row_version,
            command_idempotency_key=f"w3-{label}-apply-fixed",
        ),
    )


def test_w3_0029_current_head_postcheck_is_green(
    app_engine: Engine,
    capsys: pytest.CaptureFixture[str],
) -> None:
    with app_engine.connect() as connection:
        revision = connection.scalar(text("SELECT version_num FROM erp.alembic_version"))
        assert revision == EXPECTED_REVISION
        verify_current_0029(connection)
    print(CURRENT_0029_MARKER)
    print(HEAD_MARKER)
    output = capsys.readouterr().out
    assert CURRENT_0029_MARKER in output
    assert HEAD_MARKER in output


def test_w3_0029_http_real_workbook_stays_review_pending_and_duplicate_safe(
    app_engine: Engine,
    owner_engine: Engine,
    settings: Settings,
    seeded: SeededCase,
) -> None:
    factory = _factory(app_engine)
    application = create_app()

    def override_session() -> Iterator[Session]:
        session = factory()
        try:
            yield session
        finally:
            session.rollback()
            session.close()

    application.dependency_overrides[get_db_session] = override_session
    application.dependency_overrides[get_current_account] = lambda: seeded.account
    application.dependency_overrides[require_csrf] = lambda: seeded.account
    application.dependency_overrides[get_settings] = lambda: settings

    workbook = FIXTURE.read_bytes()
    with TestClient(application) as client:
        response = client.post(
            "/api/v1/w3/import-runs",
            data={"source_type": "RFID", "target_date": TARGET_DATE.isoformat()},
            files={"file": (FIXTURE.name, workbook, XLSX_MEDIA_TYPE)},
        )
        assert response.status_code == 201, response.text
        body = response.json()
        run = body["latest_run"]
        assert run["status"] == "PREVIEW_READY"
        assert run["counts"] == {
            "raw_rows": 314,
            "normalized_rows": 314,
            "target_rows": 36,
            "derived_groups": 0,
            "auto_matches": 0,
            "manual_matches": 0,
            "review_pending": 36,
            "blocked": 0,
        }
        assert run["warning_codes"] == ["EXPORT_CONTAINS_OTHER_DATES"]
        assert run["can_confirm"] is False
        serialized = json.dumps(body, ensure_ascii=False)
        for forbidden in (
            "storage_locator",
            "content_digest",
            "content_bytes",
            "staff_external_number",
            "legacy_staff_key",
            "mobile_phone",
        ):
            assert forbidden not in serialized

        confirm = client.post(
            f"/api/v1/w3/import-runs/{run['id']}/confirm",
            json={
                "expected_row_version": run["row_version"],
                "preview_digest": run["preview_digest"],
                "command_idempotency_key": "w3-http-pending-confirm",
            },
        )
        assert confirm.status_code == 422, confirm.text
        error = confirm.json()
        assert error["error"]["code"] == "W3_REVIEW_PENDING"
        assert error["details"] == {
            "blocked": 0,
            "review_pending": 36,
            "business_write_count": 0,
        }

        duplicate = client.post(
            "/api/v1/w3/import-runs",
            data={"source_type": "RFID", "target_date": TARGET_DATE.isoformat()},
            files={"file": ("same-bytes-again.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )
        assert duplicate.status_code == 201, duplicate.text
        assert duplicate.json()["latest_run"]["id"] == run["id"]

    with owner_engine.connect() as connection:
        counts = connection.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM erp.w3_source_receipt WHERE snapshot_id = r.snapshot_id),
                    (SELECT count(*) FROM erp.w3_import_attempt WHERE import_run_id = r.id),
                    (SELECT count(*) FROM erp.w3_normalized_rfid_row WHERE import_run_id = r.id),
                    (SELECT count(*) FROM erp.w3_actual_work_revision),
                    (SELECT count(*) FROM erp.w3_apply_control)
                  FROM erp.w3_import_run AS r
                 WHERE r.id = :run_id
                """
            ),
            {"run_id": run["id"]},
        ).one()
    assert tuple(map(int, counts)) == (2, 2, 314, 0, 0)


def test_w3_0029_command_key_lock_is_global_across_transactions(
    app_engine: Engine,
    superuser_engine: Engine,
    settings: Settings,
) -> None:
    factory = _factory(app_engine)
    holder_has_lock = Event()
    contender_pid_ready = Event()
    contender_pid: list[int] = []

    def hold_command_key() -> bool:
        with factory() as session:
            W3Service(session, settings)._lock_command_key(
                "manual-supplement",
                "w3-global-command-lock-proof",
            )
            holder_has_lock.set()
            assert contender_pid_ready.wait(timeout=10)
            deadline = time.monotonic() + 5
            observed_wait = False
            while time.monotonic() < deadline:
                with superuser_engine.connect() as observer:
                    observed_wait = bool(
                        observer.scalar(
                            text(
                                """
                                SELECT wait_event_type = 'Lock'
                                  FROM pg_stat_activity
                                 WHERE pid = :pid
                                """
                            ),
                            {"pid": contender_pid[0]},
                        )
                    )
                if observed_wait:
                    break
                time.sleep(0.01)
            session.commit()
            return observed_wait

    def contend_for_command_key() -> None:
        with factory() as session:
            assert holder_has_lock.wait(timeout=10)
            pid = session.scalar(text("SELECT pg_backend_pid()"))
            assert pid is not None
            contender_pid.append(int(pid))
            contender_pid_ready.set()
            W3Service(session, settings)._lock_command_key(
                "manual-supplement",
                "w3-global-command-lock-proof",
            )
            session.commit()

    with ThreadPoolExecutor(max_workers=2) as executor:
        holder = executor.submit(hold_command_key)
        contender = executor.submit(contend_for_command_key)
        assert holder.result(timeout=20)
        contender.result(timeout=20)


def test_w3_0029_atomic_apply_supplement_plan_and_lineage(
    app_engine: Engine,
    owner_engine: Engine,
    superuser_engine: Engine,
    settings: Settings,
    seeded: SeededCase,
) -> None:
    factory = _factory(app_engine)
    with factory() as session:
        wrong_date_link = W3MatchingRepository(session).validated_manual_link(
            service_date=TARGET_DATE,
            recipient_certification_number=START_ONLY_CERTIFICATION_NUMBER,
            service_category="방문요양",
            staff_external_number=None,
            planned_start=None,
            planned_end=None,
            recipient_id=seeded.recipient_id,
            certification_period_id=seeded.certification_period_id,
            staff_id=seeded.workers[0].staff_id,
            employment_id=seeded.workers[0].employment_id,
            service_type_id=seeded.service_type_id,
            recipient_contract_id=seeded.recipient_contract_id,
            care_assignment_id=seeded.workers[0].care_assignment_id,
            w2_schedule_id=seeded.wrong_date_schedule_id,
        )
    assert wrong_date_link is None

    prepared = (
        _prepare_confirmed_run(
            factory=factory,
            settings=settings,
            seeded=seeded,
            worker=seeded.workers[0],
            content=_single_start_only_workbook(title="candidate-A"),
            label="candidate-a",
        ),
        _prepare_confirmed_run(
            factory=factory,
            settings=settings,
            seeded=seeded,
            worker=seeded.workers[1],
            content=_single_start_only_workbook(title="candidate-B"),
            label="candidate-b",
        ),
    )
    typed_revalidation_run = _prepare_confirmed_run(
        factory=factory,
        settings=settings,
        seeded=seeded,
        worker=seeded.workers[0],
        content=_single_start_only_workbook(title="typed-revalidation"),
        label="typed-revalidation",
    )
    finalized_apply_run = _prepare_confirmed_run(
        factory=factory,
        settings=settings,
        seeded=seeded,
        worker=seeded.workers[1],
        content=_single_start_only_workbook(title="finalized-apply"),
        label="finalized-apply",
    )

    with owner_engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO erp.w3_apply_control (
                    source_type, target_date, active_snapshot_id,
                    active_import_run_id, row_version, updated_by_account_id
                ) VALUES ('RFID', :target_date, NULL, NULL, 1, :account_id)
                """
            ),
            {"target_date": TARGET_DATE, "account_id": seeded.account.id},
        )

    holder_has_lock = Event()
    contender_pid_ready = Event()
    contender_pid: list[int] = []

    def apply_contender() -> str:
        with factory() as session:
            assert holder_has_lock.wait(timeout=10)
            pid = session.scalar(text("SELECT pg_backend_pid()"))
            assert pid is not None
            contender_pid.append(int(pid))
            contender_pid_ready.set()
            result = W3Service(session, settings).apply_run(
                prepared[0].run_id,
                prepared[0].apply_request,
                seeded.account,
            )
            target = next(
                item for item in result.recent_runs if item.id == prepared[0].run_id
            )
            return target.status.value

    def apply_lock_holder() -> tuple[str, bool]:
        with factory() as session:
            session.execute(
                text(
                    """
                    SELECT row_version
                      FROM erp.w3_apply_control
                     WHERE source_type = 'RFID' AND target_date = :target_date
                     FOR UPDATE
                    """
                ),
                {"target_date": TARGET_DATE},
            ).one()
            holder_has_lock.set()
            assert contender_pid_ready.wait(timeout=10)
            deadline = time.monotonic() + 5
            observed_wait = False
            while time.monotonic() < deadline:
                with superuser_engine.connect() as observer:
                    observed_wait = bool(
                        observer.scalar(
                            text(
                                """
                                SELECT wait_event_type = 'Lock'
                                  FROM pg_stat_activity
                                 WHERE pid = :pid
                                """
                            ),
                            {"pid": contender_pid[0]},
                        )
                    )
                if observed_wait:
                    break
                time.sleep(0.01)
            result = W3Service(session, settings).apply_run(
                prepared[1].run_id,
                prepared[1].apply_request,
                seeded.account,
            )
            target = next(
                item for item in result.recent_runs if item.id == prepared[1].run_id
            )
            return target.status.value, observed_wait

    with ThreadPoolExecutor(max_workers=2) as executor:
        holder_future = executor.submit(apply_lock_holder)
        contender_future = executor.submit(apply_contender)
        holder_status, observed_wait = holder_future.result(timeout=20)
        contender_status = contender_future.result(timeout=20)
    assert holder_status == contender_status == "APPLIED"
    assert observed_wait, "the second APPLY never waited on the source/date control row"

    with owner_engine.connect() as connection:
        revisions = connection.execute(
            text(
                """
                SELECT id, import_run_id, match_decision_id, staff_id, w2_schedule_id,
                       prior_revision_id, superseded_at_utc, source_event_state,
                       actual_start
                  FROM erp.w3_actual_work_revision
                 ORDER BY id
                """
            )
        ).mappings().all()
        control = connection.execute(
            text(
                """
                SELECT active_snapshot_id, active_import_run_id, row_version
                  FROM erp.w3_apply_control
                 WHERE source_type = 'RFID' AND target_date = :target_date
                """
            ),
            {"target_date": TARGET_DATE},
        ).mappings().one()
        snapshot_states = connection.execute(
            text(
                """
                SELECT status, count(*)
                  FROM erp.w3_source_snapshot
                 WHERE source_type = 'RFID' AND target_date = :target_date
                   AND id IN (
                       SELECT snapshot_id FROM erp.w3_import_run WHERE id = ANY(:run_ids)
                   )
                 GROUP BY status
                 ORDER BY status
                """
            ),
            {
                "target_date": TARGET_DATE,
                "run_ids": [item.run_id for item in prepared],
            },
        ).all()
    assert len(revisions) == 2
    active = next(row for row in revisions if row["superseded_at_utc"] is None)
    prior = next(row for row in revisions if row["superseded_at_utc"] is not None)
    assert active["prior_revision_id"] == prior["id"]
    assert int(active["staff_id"]) == seeded.workers[0].staff_id
    assert int(prior["staff_id"]) == seeded.workers[1].staff_id
    assert active["source_event_state"] == prior["source_event_state"] == "START_ONLY"
    assert int(control["active_import_run_id"]) == prepared[0].run_id
    assert int(control["row_version"]) == 3
    assert snapshot_states == [("ACTIVE", 1), ("SUPERSEDED", 1)]

    for prepared_run in prepared:
        with factory() as session:
            retried = W3Service(session, settings).apply_run(
                prepared_run.run_id,
                prepared_run.apply_request,
                seeded.account,
            )
        retried_run = next(
            item for item in retried.recent_runs if item.id == prepared_run.run_id
        )
        assert retried_run.status.value == "APPLIED"
        with factory() as session:
            service = W3Service(session, settings)

            def conflicting_apply(
                service: W3Service = service,
                prepared_run: PreparedRun = prepared_run,
            ) -> object:
                return service.apply_run(
                    prepared_run.run_id,
                    prepared_run.apply_request.model_copy(
                        update={
                            "expected_row_version": (
                                prepared_run.apply_request.expected_row_version + 1
                            )
                        }
                    ),
                    seeded.account,
                )

            apply_conflict = _domain_code(
                conflicting_apply,
                "W3_IDEMPOTENCY_CONFLICT",
                409,
            )
            assert apply_conflict.details["entity"] == "w3_import_run"
            assert "latest" in apply_conflict.details

    with owner_engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE erp.care_assignment
                   SET invalidated_at_utc = now(),
                       updated_at_utc = now(),
                       updated_by_account_id = :account_id,
                       row_version = row_version + 1
                 WHERE id = :assignment_id
                """
            ),
            {
                "account_id": seeded.account.id,
                "assignment_id": seeded.workers[0].care_assignment_id,
            },
        )
    try:
        with factory() as session:
            service = W3Service(session, settings)
            typed_error = _domain_code(
                lambda: service.apply_run(
                    typed_revalidation_run.run_id,
                    typed_revalidation_run.apply_request,
                    seeded.account,
                ),
                "W3_TYPED_LINK_INVALID",
                422,
            )
        assert typed_error.details == {"business_write_count": 0}
        with owner_engine.connect() as connection:
            typed_state = connection.execute(
                text(
                    """
                    SELECT
                        (SELECT count(*) FROM erp.w3_actual_work_revision),
                        (SELECT row_version FROM erp.w3_apply_control
                          WHERE source_type = 'RFID' AND target_date = :target_date),
                        (SELECT count(*) FROM erp.w3_import_run
                          WHERE id = :run_id AND status = 'CONFIRMED')
                    """
                ),
                {
                    "target_date": TARGET_DATE,
                    "run_id": typed_revalidation_run.run_id,
                },
            ).one()
        assert tuple(map(int, typed_state)) == (2, 3, 1)
    finally:
        with owner_engine.begin() as connection:
            connection.execute(
                text(
                    """
                    UPDATE erp.care_assignment
                       SET invalidated_at_utc = NULL,
                           updated_at_utc = now(),
                           updated_by_account_id = :account_id,
                           row_version = row_version + 1
                     WHERE id = :assignment_id
                    """
                ),
                {
                    "account_id": seeded.account.id,
                    "assignment_id": seeded.workers[0].care_assignment_id,
                },
            )

    supplement_request = W3SupplementRequest(
        action=W3SupplementAction.CREATE,
        expected_row_version=0,
        proposed_actual_end=active["actual_start"] + timedelta(hours=1),
        reason="가명 시작전송 종료 보완",
        command_idempotency_key="w3-supplement-create-fixed",
    )
    supplement_gate = Barrier(2)

    def create_same_supplement() -> W3SupplementResponse:
        with factory() as session:
            supplement_gate.wait(timeout=10)
            return W3Service(session, settings).create_supplement(
                int(active["id"]), supplement_request, seeded.account
            )

    with ThreadPoolExecutor(max_workers=2) as executor:
        supplement_futures = [
            executor.submit(create_same_supplement),
            executor.submit(create_same_supplement),
        ]
        supplement_results = [future.result(timeout=20) for future in supplement_futures]
    supplement = supplement_results[0]
    assert supplement_results[1] == supplement
    assert supplement.row_version == 1
    assert supplement.proposed_actual_end == supplement_request.proposed_actual_end
    with factory() as session:
        service = W3Service(session, settings)
        supplement_conflict = _domain_code(
            lambda: service.create_supplement(
                int(active["id"]),
                supplement_request.model_copy(update={"reason": "다른 보완 근거"}),
                seeded.account,
            ),
            "W3_IDEMPOTENCY_CONFLICT",
            409,
        )
    assert supplement_conflict.details["entity"] == "w3_manual_supplement"
    assert supplement_conflict.details["current_row_version"] == 1
    assert "latest" in supplement_conflict.details
    with factory() as session:
        service = W3Service(session, settings)
        invalid_transition = _domain_code(
            lambda: service.create_supplement(
                int(active["id"]),
                supplement_request.model_copy(
                    update={
                        "expected_row_version": 1,
                        "command_idempotency_key": "w3-supplement-invalid-create",
                    }
                ),
                seeded.account,
            ),
            "W3_RUN_STATE_INVALID",
            422,
        )
    assert invalid_transition.details == {
        "reason": "INVALID_SUPPLEMENT_TRANSITION",
        "business_write_count": 0,
    }

    with owner_engine.connect() as connection:
        versions = connection.execute(
            text(
                """
                SELECT s.row_version AS schedule_version,
                       c.row_version AS month_version
                  FROM erp.w2_schedule AS s
                  JOIN erp.w2_schedule_month_control AS c
                    ON c.schedule_month = s.schedule_month
                 WHERE s.id = :schedule_id
                """
            ),
            {"schedule_id": int(active["w2_schedule_id"])},
        ).mappings().one()
    plan_request = W3PlanAdjustmentRequest(
        expected_schedule_row_version=int(versions["schedule_version"]),
        expected_month_row_version=int(versions["month_version"]),
        rule_version="w3-rfid-adjustment-v1",
        reason="가명 실제시간 최소오차 정렬",
        command_idempotency_key="w3-plan-adjust-fixed",
    )
    plan_gate = Barrier(2)

    def adopt_same_plan() -> W3PlanAdjustmentResponse:
        with factory() as session:
            plan_gate.wait(timeout=10)
            return W3Service(session, settings).adopt_plan_adjustment(
                int(active["id"]), plan_request, seeded.account
            )

    with ThreadPoolExecutor(max_workers=2) as executor:
        plan_futures = [
            executor.submit(adopt_same_plan),
            executor.submit(adopt_same_plan),
        ]
        plan_results = [future.result(timeout=20) for future in plan_futures]
    plan = plan_results[0]
    assert plan_results[1] == plan
    assert plan.schedule_row_version == plan_request.expected_schedule_row_version + 1
    assert plan.month_row_version == plan_request.expected_month_row_version + 1
    assert plan.adopted_planned_start == datetime(2026, 7, 6, 2, 25, tzinfo=UTC)
    assert plan.adopted_planned_end == datetime(2026, 7, 6, 3, 25, tzinfo=UTC)
    with factory() as session:
        service = W3Service(session, settings)
        plan_conflict = _domain_code(
            lambda: service.adopt_plan_adjustment(
                int(active["id"]),
                plan_request.model_copy(
                    update={
                        "expected_schedule_row_version": plan.schedule_row_version
                    }
                ),
                seeded.account,
            ),
            "W3_IDEMPOTENCY_CONFLICT",
            409,
        )
    assert plan_conflict.details["entity"] == "w3_plan_adjustment"
    assert plan_conflict.details["current_row_version"] == plan.schedule_row_version
    assert "latest" in plan_conflict.details

    with owner_engine.begin() as connection:
        connection.execute(
            text(
                """
                UPDATE erp.w2_schedule_month_control
                   SET finalized_at_utc = now(),
                       finalized_by_account_id = :account_id
                 WHERE schedule_month = :month
                """
            ),
            {"account_id": seeded.account.id, "month": SCHEDULE_MONTH},
        )
    with factory() as session:
        finalized_supplement_retry = W3Service(session, settings).create_supplement(
            int(active["id"]), supplement_request, seeded.account
        )
    assert finalized_supplement_retry == supplement
    with factory() as session:
        finalized_plan_retry = W3Service(session, settings).adopt_plan_adjustment(
            int(active["id"]), plan_request, seeded.account
        )
    assert finalized_plan_retry == plan

    with factory() as session:
        service = W3Service(session, settings)
        finalized_supplement_error = _domain_code(
            lambda: service.create_supplement(
                int(active["id"]),
                W3SupplementRequest(
                    action=W3SupplementAction.REPLACE,
                    expected_row_version=1,
                    proposed_actual_end=active["actual_start"] + timedelta(minutes=65),
                    reason="확정월에서는 차단",
                    command_idempotency_key="w3-finalized-supplement",
                ),
                seeded.account,
            ),
            "W3_MONTH_FINALIZED",
            423,
        )
    assert finalized_supplement_error.details == {"business_write_count": 0}

    with factory() as session:
        service = W3Service(session, settings)
        finalized_apply_error = _domain_code(
            lambda: service.apply_run(
                finalized_apply_run.run_id,
                finalized_apply_run.apply_request,
                seeded.account,
            ),
            "W3_MONTH_FINALIZED",
            423,
        )
    assert finalized_apply_error.details == {"business_write_count": 0}

    with owner_engine.connect() as connection:
        counts = connection.execute(
            text(
                """
                SELECT
                    (SELECT count(*) FROM erp.w3_actual_work_revision),
                    (SELECT count(*) FROM erp.w3_manual_supplement_event),
                    (SELECT count(*) FROM erp.w3_plan_adjustment_event),
                    (SELECT count(*) FROM erp.audit_event
                      WHERE action_code = 'W3_PLAN_ADJUSTMENT_ADOPT'),
                    (SELECT row_version FROM erp.w3_apply_control
                      WHERE source_type = 'RFID' AND target_date = :target_date),
                    (SELECT count(*) FROM erp.w3_import_run
                      WHERE id = :run_id AND status = 'CONFIRMED')
                """
            ),
            {
                "target_date": TARGET_DATE,
                "run_id": finalized_apply_run.run_id,
            },
        ).one()
    assert tuple(map(int, counts)) == (2, 1, 1, 1, 3, 1)

    other_decision_id = int(prior["match_decision_id"])
    connection = owner_engine.connect()
    transaction = connection.begin()
    try:
        with pytest.raises(IntegrityError) as mismatch:
            connection.execute(
                text(
                    """
                    INSERT INTO erp.w3_actual_work_revision (
                        source_type, target_date, snapshot_id, import_run_id,
                        normalized_rfid_row_id, match_decision_id,
                        source_occurrence_identity, occurrence_signature,
                        occurrence_ordinal, recipient_id, certification_period_id,
                        staff_id, employment_id, service_type_id,
                        recipient_contract_id, care_assignment_id, w2_schedule_id,
                        source_event_state, reference_minutes, actual_start,
                        actual_end, actual_seconds, fact_digest, prior_revision_id,
                        superseded_at_utc, created_by_account_id
                    )
                    SELECT source_type, target_date, snapshot_id, import_run_id,
                           normalized_rfid_row_id, :other_decision_id,
                           source_occurrence_identity, occurrence_signature,
                           occurrence_ordinal, recipient_id, certification_period_id,
                           staff_id, employment_id, service_type_id,
                           recipient_contract_id, care_assignment_id, w2_schedule_id,
                           source_event_state, reference_minutes, actual_start,
                           actual_end, actual_seconds, fact_digest, prior_revision_id,
                           now(), created_by_account_id
                      FROM erp.w3_actual_work_revision
                     WHERE id = :active_id
                    """
                ),
                {
                    "other_decision_id": other_decision_id,
                    "active_id": int(active["id"]),
                },
            )
        assert _sqlstate(mismatch.value) == "23503"
    finally:
        transaction.rollback()
        connection.close()


def test_w3_0029_app_and_backup_acl_are_fail_closed(
    app_engine: Engine,
    owner_engine: Engine,
) -> None:
    with app_engine.connect() as connection:
        transaction = connection.begin()
        try:
            with pytest.raises(ProgrammingError) as update_error:
                connection.execute(
                    text(
                        "UPDATE erp.w3_match_decision "
                        "SET reason_code = 'HOSTILE' WHERE false"
                    )
                )
            assert _sqlstate(update_error.value) == "42501"
        finally:
            transaction.rollback()

    with app_engine.connect() as connection:
        transaction = connection.begin()
        try:
            with pytest.raises(ProgrammingError) as delete_error:
                connection.execute(text("DELETE FROM erp.w3_actual_work_revision WHERE false"))
            assert _sqlstate(delete_error.value) == "42501"
        finally:
            transaction.rollback()

    with owner_engine.connect() as connection:
        privileges = connection.execute(
            text(
                """
                SELECT
                    has_table_privilege('erp_backup', 'erp.w3_match_decision', 'SELECT'),
                    has_table_privilege('erp_backup', 'erp.w3_match_decision', 'INSERT'),
                    has_table_privilege('erp_backup', 'erp.w3_match_decision', 'UPDATE'),
                    has_table_privilege('erp_backup', 'erp.w3_match_decision', 'DELETE')
                """
            )
        ).one()
    assert tuple(bool(value) for value in privileges) == (True, False, False, False)
